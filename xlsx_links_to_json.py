#!/usr/bin/env python3
import sys, json, unicodedata, re
from openpyxl import load_workbook

# robust slug to match your exercise ids (e.g., "Stability Ball Russian Twist" -> "stability-ball-russian-twist")
def slugify(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9]+", "-", s.strip().lower())
    return re.sub(r"-{2,}", "-", s).strip("-")

def find_header(ws):
    # scan first ~50 rows to find the header row that contains both columns
    wanted = {"exercise", "youtube"}
    for r in range(1, min(ws.max_row, 50) + 1):
        cells = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[r]]
        if wanted.issubset(set(cells)):
            headers = {cells[i]: i+1 for i in range(len(cells))}
            return r, headers
    raise SystemExit("Could not find a header row with 'Exercise' and 'YouTube'.")

def extract(path, sheet_name=None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_row, headers = find_header(ws)

    name_col = headers["exercise"]
    link_col = headers["youtube"]

    out = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(r, name_col)
        link_cell = ws.cell(r, link_col)

        name = (name_cell.value or "").strip()
        if not name:
            continue

        url = None
        # hyperlinks added via UI
        if link_cell.hyperlink and link_cell.hyperlink.target:
            url = link_cell.hyperlink.target

        # hyperlinks created via =HYPERLINK("url","text")
        if not url and isinstance(link_cell.value, str) and link_cell.value.startswith("="):
            m = re.search(r'HYPERLINK\("([^"]+)"', link_cell.value, flags=re.I)
            if m:
                url = m.group(1)

        if not url:
            continue

        out[slugify(name)] = {"exercise": name, "youtube": url}

    return out

def main():
    if len(sys.argv) < 2:
        print("Usage: xlsx_links_to_json.py <input.xlsx> [output.json]")
        sys.exit(1)
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else "exercise_media.json"
    data = extract(src)
    with open(dst, "w", encoding="utf-8") as f:
        json.dump([{"id": k, **v} for k,v in data.items()], f, indent=2, ensure_ascii=False)
    print(f"Wrote {len(data)} items to {dst}")

if __name__ == "__main__":
    main()